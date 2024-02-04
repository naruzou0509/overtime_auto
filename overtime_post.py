# module
import os
import datetime
from time import sleep
import openpyxl as op
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
import overtime_function as of

# 変数
user = os.getlogin() # PCのログインユーザー名取得
url = input("URLを入力してください。：") # 対象のサイト
id_tp = input("ログインIDを入力してください。：") # ログインID
pw_tp = input("パスワードを入力してください。：") # パスワード
options = webdriver.ChromeOptions() # seleniumのオプション
options.add_experimental_option("detach", True) # seleniumにオプション追加(処理終了後でも画面を開いたままに)
driver = webdriver.Chrome(options=options) # クロームドライバー（新バージョンではドライバーのDL不要）

# サイト表示
driver.get(url)

#カレントウインドウを最大化する
driver.maximize_window()
sleep(3)

# ログイン処理
ele_r = driver.find_element(By.NAME, "email") # ログインIDの場所特定
ele_r.send_keys(id_tp) # ログインIDの入力

ele_p = driver.find_element(By.NAME, "password") # パスワードの場所特定
ele_p.send_keys(pw_tp) # パスワードの入力

ele_rb = driver.find_element(By.CLASS_NAME, "btn-block") # ログインボタンの場所特定
ele_rb.click() # ログインボタンのクリック
sleep(3)

# 「リシテア」を表示
ele_sk = driver.find_element(By.CLASS_NAME, "btn-secondary") # スキップボタンの場所特定
ele_sk.click() # ログインボタンのクリック
sleep(1)

ele_rk = driver.find_elements(By.CSS_SELECTOR, ".col-xs-12.col-sm-4.col-md-3") # スキップボタンの場所特定
ele_rk[1].click() # スキップボタンのクリック
sleep(5)

# 「残業申請」を表示
handle_array = driver.window_handles # ウィンドウハンドルを取得する
driver.switch_to.window(handle_array[1]) # seleniumで操作可能なdriverを切り替える
driver.execute_script("return menujs.go('WC230','PERSONAL')") # 残業申請をクリック
sleep(3)

# beutifulsoup準備
html = driver.page_source.encode('utf-8')
soup = BeautifulSoup(html, 'html.parser')

# 現在表示されている月の取得
ele_mon = soup.select("#WC230 > form:nth-child(1) > header > nav.funcNav > div.container > div > div > div:nth-child(1) > div > div.date > div > div > span:nth-child(3)") # 該当月のタグ取得
ele_mon_data = int(ele_mon[0].contents[0]) # 該当月のタグの値取得
print(type(ele_mon_data)) # 型の確認
print(str(ele_mon_data) + "月")

# Excelの対象月の値の取得
cell_place = "I2" # cellの場所
cell_mon = of.wb_value(user,cell_place) # 対象月の取得
print(type(cell_mon)) # 型の確認
print(str(cell_mon) + "月")

# 取得した月の値とExcelの対象月の値の比較による分岐
if ele_mon_data == cell_mon :
    print("月は等しいのでそのまま処理をします。")
    sleep(3)
elif ele_mon_data > cell_mon :
    print("現在表示されている月がExcelの値より大きいので「←」をクリックします。")
    driver.execute_script("return doLabel('PREVIOUSMONTH');") # 「←」をクリック
    sleep(3)
elif ele_mon_data < cell_mon :
    print("現在表示されている月がExcelの値より小さいので「→」をクリックします。")
    driver.execute_script("return doLabel('NEXTMONTH');") # 「→」をクリック
    sleep(3)
else :
    print("該当なし")

# Excelの月末の値の取得
cell_place = "K2" # cellの場所
cell_eom = of.wb_value(user,cell_place) # 月末の取得
print(type(cell_eom)) # 型の確認
print("月末は"+str(cell_eom)+"日")

# 登録処理
for i in range(cell_eom): # 繰り返し処理
    num = i + 2 # cellの場所の値
    print(num)
    wb = op.load_workbook(r"C:\Users\{}\Documents\overtime_list_filter.xlsm".format(user), data_only=True) # book読み込み
    ws = wb["List"] # シートの「List」を指定
    cell_day = ws["A"+ str(num)].value # 登録の値取得
    print(cell_day)
    cell_day_format = cell_day.date().strftime('%Y/%m/%d')
    print(cell_day_format)
    cell_reg = ws["E"+ str(num)].value # 登録の値取得
    print(cell_reg)
    wb.close() # bookを閉じる

    if not cell_reg : # 変数が空のとき
        continue

    driver.execute_script("return doLabelYearMonthDay('YEARMONTHDAY','{}')".format(cell_day_format)) # 日付クリック
    sleep(3)
    dropdown = driver.find_element(By.NAME, "TimecardCause") # ドロップダウンリストを選択
    select = Select(dropdown)
    select.select_by_value("Cause_6") # valueが"Cause_6"のoptionタグを選択状態にする
    sleep(3)
    driver.execute_script("top.dosubmitRegister();return false;") # 登録クリック
    sleep(3)
    driver.execute_script("return top.doLabel('BACK')") # 戻るクリック
    sleep(3)

sleep(10000)

# 転記処理
cell_place_list_t =["B","C","D"] # cellの頭の文字のリスト

for i in range(cell_eom): # 繰り返し処理
    num = i + 2 # cellの場所の値
    cell_place_list_ut = [] # listの初期化

    for e in range(3): # 繰り返しでlistに格納
        cell_place_list_ut.append(cell_place_list_t[e]+str(num))# listに追加

    of.input_post(user,cell_place_list_ut,driver,i) # 転記

print("処理が完了しました。")
